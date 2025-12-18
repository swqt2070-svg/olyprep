from fastapi import (
    APIRouter,
    Depends,
    Request,
    Form,
    status,
    HTTPException,
    UploadFile,
    File,
)
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, Response
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from sqlalchemy import func, or_, not_
from typing import List, Optional
import json
import io
import zipfile
import re
import json
from pathlib import Path
from uuid import uuid4
import os
import csv

from app.deps import get_db, get_current_user, require_role, require_teacher_or_admin
from app.models import (
    User,
    Question,
    Test,
    TestQuestion,
    Submission,
    Answer,
    AnswerOption,
    Category,
    RegistrationCode,
    UserRole,
)
from app.security import hash_password, verify_password, create_token

router = APIRouter(prefix="/ui", tags=["ui"])
templates = Jinja2Templates(directory="app/templates")

STUDENT_INVITE_CODE = "STUDENT2025"
TEACHER_INVITE_CODE = "TEACHER2025"
BASE_DIR = Path(__file__).resolve().parent.parent
UPLOAD_DIR = BASE_DIR / "static" / "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

GRADE_CHOICES = ["9", "10", "11"]
STAGE_CHOICES = ["Школьный", "Муниципальный", "Региональный", "Заключительный"]


def redirect(url: str) -> RedirectResponse:
    return RedirectResponse(url=url, status_code=status.HTTP_303_SEE_OTHER)


# ---------- CATEGORIES HELPERS ----------


def _fetch_categories(db: Session) -> List[Category]:
    return (
        db.query(Category)
        .order_by(Category.parent_id.asc(), Category.name.asc())
        .all()
    )


def _build_category_tree(categories: List[Category]) -> List[dict]:
    by_parent: dict[Optional[int], list[Category]] = {}
    for c in categories:
        by_parent.setdefault(c.parent_id, []).append(c)
    for lst in by_parent.values():
        lst.sort(key=lambda x: (x.name or "").lower())

    def walk(pid: Optional[int] = None) -> List[dict]:
        items: list[dict] = []
        for c in by_parent.get(pid, []):
            items.append({"cat": c, "children": walk(c.id)})
        return items

    return walk(None)


def _build_category_choices(categories: List[Category], roots_only: bool = False) -> List[dict]:
    by_parent: dict[Optional[int], list[Category]] = {}
    for c in categories:
        by_parent.setdefault(c.parent_id, []).append(c)
    for lst in by_parent.values():
        lst.sort(key=lambda x: (x.name or "").lower())

    result: list[dict] = []

    def walk(pid: Optional[int], prefix: str = "", depth: int = 0) -> None:
        for c in by_parent.get(pid, []):
            path_label = ""
            try:
                path_label = c.full_path
            except Exception:
                path_label = c.name
            indent = "— " * depth
            result.append({"id": c.id, "label": f"{indent}{path_label}"})
            walk(c.id, prefix + "— ", depth + 1)

    if roots_only:
        for c in by_parent.get(None, []):
            label = c.name
            result.append({"id": c.id, "label": label})
    else:
        walk(None, "")
    return result


def _get_root_category_choices(db: Session) -> List[dict]:
    return _build_category_choices(_fetch_categories(db), roots_only=True)


def _category_label(obj: Question) -> str:
    try:
        rel = getattr(obj, "category_rel", None)
        if rel:
            return rel.full_path
    except Exception:
        pass
    return getattr(obj, "category", None) or "Без категории"


# ---------- UPLOADS ----------


@router.post("/upload-image")
async def upload_image(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
):
    if not file or not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Можно загружать только изображения")

    ext = (Path(file.filename or "").suffix or "").lower()
    allowed_ext = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
    if ext and ext not in allowed_ext:
        raise HTTPException(status_code=400, detail="Недопустимое расширение файла")
    if not ext:
        ext = ".img"

    data = await file.read()
    max_size = 5 * 1024 * 1024  # 5 MB
    if len(data) > max_size:
        raise HTTPException(status_code=413, detail="Файл слишком большой")

    filename = f"{uuid4().hex}{ext}"
    dest = UPLOAD_DIR / filename
    dest.write_bytes(data)

    url = f"/static/uploads/{filename}"
    return JSONResponse({"url": url})


# ---------- ВСПОМОГАТЕЛЬНОЕ: контекст ЛК ----------


def build_account_context(
    request: Request,
    db: Session,
    user: User,
    search: str = "",
    password_error: Optional[str] = None,
    password_success: Optional[str] = None,
):
    student_results = None
    teacher_results = None

    if user.role == "student":
        submissions: List[Submission] = (
            db.query(Submission)
            .filter(Submission.user_id == user.id)
            .order_by(Submission.id.desc())
            .all()
        )
        results = []
        for sub in submissions:
            test = db.get(Test, sub.test_id)
            if not test:
                continue
            tqs = (
                db.query(TestQuestion)
                .filter(TestQuestion.test_id == test.id)
                .all()
            )
            max_points = sum(tq.points for tq in tqs) if tqs else 0
            results.append(
                {
                    "submission": sub,
                    "test": test,
                    "max_points": max_points,
                }
            )
        student_results = results

    if user.role in ("teacher", "admin"):
        students: List[User] = db.query(User).filter(User.role == "student").all()
        students_map = {s.id: s for s in students}
        student_ids = list(students_map.keys())

        if student_ids:
            submissions2: List[Submission] = (
                db.query(Submission)
                .filter(Submission.user_id.in_(student_ids))
                .order_by(Submission.id.desc())
                .all()
            )
        else:
            submissions2 = []

        tests: List[Test] = db.query(Test).all()
        tests_map = {t.id: t for t in tests}
        max_points_map: dict[int, int] = {}

        for t in tests:
            tqs = (
                db.query(TestQuestion)
                .filter(TestQuestion.test_id == t.id)
                .all()
            )
            max_points_map[t.id] = sum(tq.points for tq in tqs) if tqs else 0

        rows = []
        for sub in submissions2:
            student = students_map.get(sub.user_id)
            test = tests_map.get(sub.test_id)
            if not student or not test:
                continue
            max_points = max_points_map.get(test.id, 0)
            display_name = getattr(student, "full_name", None) or student.email
            rows.append(
                {
                    "submission": sub,
                    "student": student,
                    "test": test,
                    "max_points": max_points,
                    "display_name": display_name,
                }
            )
        if search:
            s = search.lower()
            rows = [
                r
                for r in rows
                if s in str(r["display_name"]).lower() or s in str(r["student"].email).lower()
            ]

        teacher_results = rows

    return {
        "request": request,
        "user": user,
        "search": search,
        "password_error": password_error,
        "password_success": password_success,
        "student_results": student_results,
        "teacher_results": teacher_results,
    }


# ---------- CATEGORIES UI ----------


def _categories_page_context(
    request: Request,
    db: Session,
    user: User,
    error: Optional[str] = None,
    success: Optional[str] = None,
):
    categories = _fetch_categories(db)
    return {
        "request": request,
        "user": user,
        "category_tree": _build_category_tree(categories),
        "category_choices": _build_category_choices(categories),
        "error": error,
        "success": success,
    }


@router.get("/categories", response_class=HTMLResponse)
async def categories_list(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    return templates.TemplateResponse(
        "categories.html",
        _categories_page_context(request, db, user),
    )


@router.post("/categories", response_class=HTMLResponse)
async def categories_create(
    request: Request,
    name: str = Form(...),
    parent_id: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    name = (name or "").strip()
    parent_id_int: Optional[int] = None
    if parent_id not in (None, "", "0"):
        try:
            parent_id_int = int(parent_id)
        except ValueError:
            parent_id_int = None

    parent: Optional[Category] = None
    if parent_id_int:
        parent = db.get(Category, parent_id_int)
        if not parent:
            ctx = _categories_page_context(
                request, db, user, error="Родительская категория не найдена."
            )
            return templates.TemplateResponse(
                "categories.html",
                ctx,
                status_code=400,
            )

    if not name:
        ctx = _categories_page_context(
            request, db, user, error="Название категории обязательно."
        )
        return templates.TemplateResponse(
            "categories.html",
            ctx,
            status_code=400,
        )

    cat = Category(name=name, parent_id=parent.id if parent else None)
    db.add(cat)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        ctx = _categories_page_context(
            request, db, user, error="Такая категория уже существует на этом уровне."
        )
        return templates.TemplateResponse(
            "categories.html",
            ctx,
            status_code=400,
        )

    ctx = _categories_page_context(
        request, db, user, success="Категория сохранена."
    )
    return templates.TemplateResponse("categories.html", ctx)


# ---------- AUTH UI ----------


@router.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse(
        "login.html",
        {"request": request, "user": None, "error": None},
    )


@router.post("/login")
async def login_submit(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.email == email).first()
    if not user or not verify_password(password, user.password_hash):
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "user": None, "error": "Неверная почта или пароль"},
            status_code=400,
        )
    if not getattr(user, "active", True):
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "user": None, "error": "Учётная запись заморожена. Обратитесь к администратору."},
            status_code=403,
        )

    token = create_token({"id": user.id, "role": user.role})
    response = redirect("/ui/dashboard")
    response.set_cookie("access_token", token, httponly=True)
    return response


@router.get("/register", response_class=HTMLResponse)
async def register_page(request: Request):
    # Регистрация отключена, отправляем на страницу логина с подсказкой
    return templates.TemplateResponse(
        "login.html",
        {
            "request": request,
            "user": None,
            "error": "Регистрация отключена. Обратитесь к администратору.",
        },
        status_code=400,
    )


@router.post("/register")
async def register_submit(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    invite_code: str = Form(""),
    full_name: str = Form(""),
    student_class: str = Form(""),
    db: Session = Depends(get_db),
):
    return templates.TemplateResponse(
        "login.html",
        {
            "request": request,
            "user": None,
            "error": "Регистрация отключена. Обратитесь к администратору.",
        },
        status_code=400,
    )

    email = email.strip()
    invite_code = invite_code.strip()
    full_name = full_name.strip()
    student_class = student_class.strip()

    if db.query(User).filter(User.email == email).first():
        return templates.TemplateResponse(
            "register.html",
            {
                "request": request,
                "user": None,
                "error": "Такая почта уже используется",
                "success": None,
                "full_name": full_name,
            },
            status_code=400,
        )

    has_admin = db.query(User).filter(User.role == "admin").first() is not None
    code_rec: Optional[RegistrationCode] = None

    if not has_admin:
        role = "admin"
    else:
        if not invite_code:
            return templates.TemplateResponse(
                "register.html",
                {
                    "request": request,
                    "user": None,
                    "error": "Для регистрации нужен код приглашения. Получите его у учителя или администратора.",
                    "success": None,
                    "full_name": full_name,
                },
                status_code=400,
            )
        code_rec = (
            db.query(RegistrationCode)
            .filter(RegistrationCode.code == invite_code)
            .first()
        )
        if code_rec:
            if code_rec.used >= code_rec.max_uses:
                return templates.TemplateResponse(
                    "register.html",
                    {
                    "request": request,
                    "user": None,
                    "error": "Лимит регистраций по этому коду исчерпан.",
                    "success": None,
                    "full_name": full_name,
                },
                status_code=400,
            )
            role = code_rec.role
        elif invite_code == STUDENT_INVITE_CODE:
            role = "student"
        elif invite_code == TEACHER_INVITE_CODE:
            role = "teacher"
        else:
            return templates.TemplateResponse(
                "register.html",
                {
                    "request": request,
                    "user": None,
                    "error": "Неверный код приглашения.",
                    "success": None,
                    "full_name": full_name,
                },
                status_code=400,
            )

    if not full_name:
        return templates.TemplateResponse(
            "register.html",
            {
                "request": request,
                "user": None,
                "error": "Укажите ФИО.",
                "success": None,
                "full_name": full_name,
            },
            status_code=400,
        )

    user = User(
        email=email,
        password_hash=hash_password(password),
        role=role,
        full_name=full_name,
        student_class=None,
    )
    db.add(user)
    if code_rec:
        code_rec.used = (code_rec.used or 0) + 1
        db.add(code_rec)
    db.commit()

    token = create_token({"id": user.id, "role": user.role})
    response = redirect("/ui/dashboard")
    response.set_cookie("access_token", token, httponly=True)
    return response


@router.post("/logout")
async def logout():
    response = redirect("/ui/login")
    response.delete_cookie("access_token")
    return response


# ---------- DASHBOARD ----------


@router.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request, user: User = Depends(get_current_user)):
    return redirect("/ui/account")


# ---------- ЛИЧНЫЙ КАБИНЕТ ----------


@router.get("/account", response_class=HTMLResponse)
async def account_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    search = (request.query_params.get("search") or "").strip()
    ctx = build_account_context(request, db, user, search=search)
    return templates.TemplateResponse("account.html", ctx)


@router.post("/account/change-password", response_class=HTMLResponse)
async def account_change_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    new_password2: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    error = None
    success = None

    if not verify_password(current_password, user.password_hash):
        error = "Текущий пароль введён неверно."
    elif len(new_password) < 6:
        error = "Новый пароль должен быть не короче 6 символов."
    elif new_password != new_password2:
        error = "Пароль и подтверждение не совпадают."
    else:
        user.password_hash = hash_password(new_password)
        db.add(user)
        db.commit()
        success = "Пароль успешно обновлён."

    ctx = build_account_context(request, db, user, error, success)
    return templates.TemplateResponse(
        "account.html",
        ctx,
        status_code=400 if error else 200,
    )


# ---------- ADMIN: пользователи ----------


@router.get("/admin/users", response_class=HTMLResponse)
async def admin_users_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin")),
):
    users = db.query(User).order_by(User.id.asc()).all()
    return templates.TemplateResponse(
        "users_admin.html",
        {
            "request": request,
            "user": user,
            "users": users,
            "error": None,
            "success": None,
        },
    )


@router.post("/admin/users/set-role", response_class=HTMLResponse)
async def admin_set_role(
    request: Request,
    email: str = Form(...),
    role: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin")),
):
    allowed_roles = ("admin", "teacher", "student")

    email = email.strip()
    role = role.strip()

    error: Optional[str] = None
    success: Optional[str] = None

    if role not in allowed_roles:
        error = "Недопустимая роль."
    else:
        target = db.query(User).filter(User.email == email).first()
        if not target:
            error = "Пользователь с такой почтой не найден."
        elif target.id == user.id and role != "admin":
            error = "Нельзя понизить роль собственного админ‑аккаунта."
        else:
            old_role = target.role
            target.role = role
            db.add(target)
            db.commit()
            success = f"Роль пользователя {email} изменена с {old_role} на {role}."

    users = db.query(User).order_by(User.id.asc()).all()
    status_code = 400 if error else 200
    return templates.TemplateResponse(
        "users_admin.html",
        {
            "request": request,
            "user": user,
            "users": users,
            "error": error,
            "success": success,
        },
        status_code=status_code,
    )

@router.post("/admin/users/reset-password", response_class=HTMLResponse)
async def admin_reset_password(
    request: Request,
    email: str = Form(...),
    new_password: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin")),
):
    email = email.strip()
    new_password = new_password.strip()
    error = None
    success = None
    target = db.query(User).filter(User.email == email).first()
    if not target:
        error = "Пользователь не найден."
    elif len(new_password) < 6:
        error = "Новый пароль должен быть не короче 6 символов."
    else:
        target.password_hash = hash_password(new_password)
        db.add(target)
        db.commit()
        success = f"Пароль для {email} обновлён."

    users = db.query(User).order_by(User.id.asc()).all()
    return templates.TemplateResponse(
        "users_admin.html",
        {
            "request": request,
            "user": user,
            "users": users,
            "error": error,
            "success": success,
        },
        status_code=400 if error else 200,
    )


# ---------- ADMIN: импорт пользователей из Excel/CSV ----------

def _role_from_number(num: str) -> Optional[str]:
    mapping = {"1": "student", "2": "teacher", "3": "admin"}
    return mapping.get(str(num).strip())


@router.get("/admin/users/import", response_class=HTMLResponse)
async def admin_import_users_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin")),
):
    return templates.TemplateResponse(
        "invite_codes.html",
        {
            "request": request,
            "user": user,
            "error": None,
            "success": None,
            "summary": None,
        },
    )


def _active_admins_count(db: Session) -> int:
    return (
        db.query(User)
        .filter(User.role == UserRole.ADMIN, User.active.is_(True))
        .count()
    )


@router.post("/admin/users/toggle-active", response_class=HTMLResponse)
async def admin_toggle_active(
    request: Request,
    user_id: int = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin")),
):
    target = db.get(User, user_id)
    if not target:
        return RedirectResponse("/ui/admin/users", status_code=303)
    if target.id == user.id:
        error = "Нельзя заморозить/разморозить самого себя."
    elif target.role == UserRole.ADMIN and target.active and _active_admins_count(db) <= 1:
        error = "Нельзя заморозить последнего активного админа."
    else:
        target.active = not bool(getattr(target, "active", True))
        db.add(target)
        db.commit()
        error = None
    users = db.query(User).order_by(User.id.asc()).all()
    return templates.TemplateResponse(
        "users_admin.html",
        {
            "request": request,
            "user": user,
            "users": users,
            "error": error,
            "success": None if error else "Статус учётной записи обновлён.",
        },
        status_code=400 if error else 200,
    )


@router.post("/admin/users/delete", response_class=HTMLResponse)
async def admin_delete_user(
    request: Request,
    user_id: int = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin")),
):
    target = db.get(User, user_id)
    if not target:
        return RedirectResponse("/ui/admin/users", status_code=303)
    if target.id == user.id:
        error = "Нельзя удалить самого себя."
    elif target.role == UserRole.ADMIN and _active_admins_count(db) <= 1:
        error = "Нельзя удалить последнего админа."
    else:
        db.delete(target)
        db.commit()
        error = None
    users = db.query(User).order_by(User.id.asc()).all()
    return templates.TemplateResponse(
        "users_admin.html",
        {
            "request": request,
            "user": user,
            "users": users,
            "error": error,
            "success": None if error else "Учётная запись удалена.",
        },
        status_code=400 if error else 200,
    )


@router.post("/admin/users/import", response_class=HTMLResponse)
async def admin_import_users_submit(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin")),
):
    filename = file.filename or ""
    data = await file.read()
    rows: list[list[str]] = []
    error: Optional[str] = None

    if not filename:
        error = "Файл не выбран."
    else:
        suffix = filename.lower().split(".")[-1]
        if suffix == "csv":
            try:
                text = data.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = data.decode("cp1251", errors="ignore")
            reader = csv.reader(text.splitlines())
            rows = [list(row) for row in reader]
        elif suffix in ("xlsx", "xlsm", "xltx", "xltm"):
            try:
                from openpyxl import load_workbook  # type: ignore
            except ImportError:
                error = "Для импорта .xlsx установите пакет openpyxl."
            if not error:
                wb = load_workbook(io.BytesIO(data), read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    rows.append([cell if cell is not None else "" for cell in row])
        else:
            error = "Поддерживаются файлы .csv или .xlsx"

    created = 0
    skipped = 0
    errors: list[str] = []

    if not error and rows:
        if rows and rows[0]:
            header = "".join(str(x or "").lower() for x in rows[0])
            if "роль" in header and "лог" in header:
                rows = rows[1:]

        for row in rows:
            if len(row) < 4:
                skipped += 1
                continue
            role_raw, fio, login, pwd = [str(x or "").strip() for x in row[:4]]
            role_str = _role_from_number(role_raw)
            if not role_str or not login or not pwd:
                skipped += 1
                continue
            if db.query(User).filter(User.email == login).first():
                skipped += 1
                continue
            u = User(
                email=login,
                full_name=fio or None,
                role=role_str,
                password_hash=hash_password(pwd),
            )
            db.add(u)
            created += 1
        db.commit()

    if error:
        summary = None
    else:
        summary = {"created": created, "skipped": skipped, "errors": errors}

    return templates.TemplateResponse(
        "invite_codes.html",
        {
            "request": request,
            "user": user,
            "error": error,
            "success": None if error else "Импорт завершён",
            "summary": summary,
        },
        status_code=400 if error else 200,
    )


# ---------- IMPORT: ZIP с .md ----------

def _try_parse_choice(
    question_lines: List[str],
    answer_line: str,
) -> Optional[dict]:
    """
    Пытаемся выделить варианты вида "а) текст", "б) текст" и
    понять, какой из них правильный по строке ответа.
    """
    opt_re = re.compile(r"^\s*([A-Za-zА-Яа-я])\)\s*(.+)")
    options: List[str] = []
    letter_to_index: dict[str, int] = {}
    first_opt_idx: Optional[int] = None

    for i, line in enumerate(question_lines):
        m = opt_re.match(line)
        if not m:
            continue
        letter = m.group(1).lower()
        body = m.group(2).strip()
        options.append(body)
        letter_to_index[letter] = len(options) - 1
        if first_opt_idx is None:
            first_opt_idx = i

    if len(options) < 2:
        return None

    # Ищем букву/текст в строке ответа
    ans = answer_line.strip()
    ans_letter: Optional[str] = None
    ans_clean = ""

    m_ans = opt_re.match(ans)
    if m_ans:
        ans_letter = m_ans.group(1).lower()
        ans_clean = m_ans.group(2).strip(" .;").lower()
    else:
        m_letter_only = re.search(r"([A-Za-zА-Яа-я])\)", ans)
        if m_letter_only:
            ans_letter = m_letter_only.group(1).lower()
            ans_clean = ans[m_letter_only.end() :].strip(" .;").lower()
        else:
            ans_clean = ans.strip(" .;").lower()

    correct_index: Optional[int] = None

    # Сначала пробуем по букве
    if ans_letter and ans_letter in letter_to_index:
        correct_index = letter_to_index[ans_letter]
    else:
        normalized_ans = ans_clean
        for idx, opt in enumerate(options):
            opt_norm = opt.strip(" .;").lower()
            if opt_norm and opt_norm == normalized_ans:
                correct_index = idx
                break

    if correct_index is None:
        return None

    # Текст вопроса — всё до первой строки с вариантом
    if first_opt_idx is None:
        first_opt_idx = 0
    q_lines = question_lines[:first_opt_idx]
    # обрезаем пустые
    while q_lines and not q_lines[0].strip():
        q_lines.pop(0)
    while q_lines and not q_lines[-1].strip():
        q_lines.pop()
    question_text = "\n".join(q_lines).strip()
    if not question_text:
        # на всякий случай
        question_text = "\n".join(question_lines).strip()

    return {
        "text": question_text,
        "answer_type": "single",
        "options_json": json.dumps(options, ensure_ascii=False),
        "correct": str(correct_index),
    }

def parse_markdown_to_question(raw: str) -> Optional[dict]:
    """
    Понимает файлы вида:

    **КЛАСС** #класс_9
    **ГОД** #год_1819
    **ЭТАП** #муницип
    **КАТЕГОРИЯ** [[#Дерево]]
    **ТИП ВОПРОСА** #закрытый/#открытый
    ...

    # Вопрос
    1) Текст вопроса...
    а) вариант 1;
    б) вариант 2;
    в) вариант 3;

    # Ответ
    б) вариант 2;

    Пустые шаблоны (нет текста вопроса и нет ответа) и заметки без "# Вопрос" — пропускаем.
    """
    text = raw.replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return None

    # если вообще нет "# Вопрос" – это не задача
    if not re.search(r"^\s*#\s*Вопрос\b", text, re.IGNORECASE | re.MULTILINE):
        return None

    lines = text.split("\n")

    # --- ищем строку "# Вопрос" ---
    q_header_idx = None
    for i, line in enumerate(lines):
        if re.search(r"^\s*#\s*Вопрос\b", line, re.IGNORECASE):
            q_header_idx = i
            break
    if q_header_idx is None:
        return None

    question_start = q_header_idx + 1

    # ---------- МЕТАДАННЫЕ из шапки ----------
    meta_category: Optional[str] = None
    meta_grade: Optional[int] = None
    meta_year: Optional[str] = None
    meta_stage: Optional[str] = None
    meta_type: Optional[str] = None  # 'open' / 'closed'

    header_lines = lines[:q_header_idx]

    for meta_line in header_lines:
        # КЛАСС
        m = re.match(r"\s*\*\*\s*КЛАСС\s*\*\*\s*(.+)", meta_line, re.IGNORECASE)
        if m and meta_grade is None:
            val = m.group(1)
            m2 = re.search(r"(\d+)", val)
            if m2:
                try:
                    meta_grade = int(m2.group(1))
                except ValueError:
                    pass

        # ГОД
        m = re.match(r"\s*\*\*\s*ГОД\s*\*\*\s*(.+)", meta_line, re.IGNORECASE)
        if m and meta_year is None:
            val = m.group(1)
            m2 = re.search(r"(\d+)", val)
            if m2:
                meta_year = m2.group(1)

        # ЭТАП
        m = re.match(r"\s*\*\*\s*ЭТАП\s*\*\*\s*(.+)", meta_line, re.IGNORECASE)
        if m and meta_stage is None:
            val = m.group(1).strip()
            val = re.sub(r"^#+", "", val)  # "#муницип" -> "муницип"
            meta_stage = val or None

        # КАТЕГОРИЯ
        m = re.match(r"\s*\*\*\s*КАТЕГОРИЯ\s*\*\*\s*(.+)", meta_line, re.IGNORECASE)
        if m and meta_category is None:
            val = m.group(1).strip()
            # "[[#Дерево]]" -> "Дерево"
            val = re.sub(r"^\[\[", "", val)
            val = re.sub(r"]]$", "", val)
            val = val.lstrip("#").strip()
            meta_category = val or None

        # ТИП ВОПРОСА (#закрытый / #открытый)
        m = re.match(r"\s*\*\*\s*ТИП ВОПРОСА\s*\*\*\s*(.+)", meta_line, re.IGNORECASE)
        if m and meta_type is None:
            val = m.group(1).strip().lower()
            if "закрытый" in val:
                meta_type = "closed"
            elif "открытый" in val:
                meta_type = "open"

    # ---------- границы вопроса и ответ ----------
    answer_value = ""
    question_end = len(lines)

    # 1) "Ответ: ..." в одной строке
    inline_idx = None
    for idx, line in enumerate(lines):
        m = re.search(r"(Ответ|Answer)\s*[:\-]\s*(.+)", line, re.IGNORECASE)
        if m:
            inline_idx = idx
            answer_value = m.group(2).strip()
            question_end = idx
            break

    # 2) заголовок "# Ответ"
    if inline_idx is None:
        answer_header_idx = None
        for idx, line in enumerate(lines):
            if re.search(r"^\s*#\s*Ответ\b", line, re.IGNORECASE):
                answer_header_idx = idx
                break
        if answer_header_idx is not None:
            question_end = answer_header_idx
            for j in range(answer_header_idx + 1, len(lines)):
                candidate = lines[j].strip()
                if candidate:
                    answer_value = candidate
                    break

    question_lines = lines[question_start:question_end]

    # пустой вопрос + пустой ответ -> считаем шаблоном, не импортируем
    if (not any(l.strip() for l in question_lines)) and not answer_value.strip():
        return None

    # ---------- картинка ----------
    image_name = None
    m1 = re.search(r"!\[\[(.+?)\]\]", text)
    if m1:
        image_name = m1.group(1).strip()
    else:
        m2 = re.search(r"!\[[^\]]*]\((.+?)\)", text)
        if m2:
            image_name = m2.group(1).strip()

    # ---------- выбор одного варианта ----------
    choice_data = None
    if meta_type != "open":  # для открытых сразу идём в текстовый режим
        choice_data = _try_parse_choice(question_lines, answer_value)

    if choice_data:
        # choice_data уже содержит text / answer_type / options_json / correct
        choice_data["image_name"] = image_name
        choice_data["category"] = meta_category
        choice_data["grade"] = meta_grade
        choice_data["year"] = meta_year
        choice_data["stage"] = meta_stage
        return choice_data

    # ---------- fallback: текстовый ответ ----------
    question_text = "\n".join(question_lines).strip() or text
    question_text = re.sub(r"!\[\[.+?\]\]", "", question_text)
    question_text = re.sub(r"!\[[^\]]*]\(.+?\)", "", question_text)
    question_text = question_text.strip()

    return {
        "text": question_text,
        "answer_type": "text",
        "correct": answer_value.strip() if answer_value else "",
        "options_json": None,
        "image_name": image_name,
        "category": meta_category,
        "grade": meta_grade,
        "year": meta_year,
        "stage": meta_stage,
    }


@router.get("/import", response_class=HTMLResponse)
async def import_page(
    request: Request,
    user: User = Depends(require_role("admin", "teacher")),
):
    return templates.TemplateResponse(
        "import.html",
        {
            "request": request,
            "user": user,
            "error": None,
            "summary": None,
        },
    )


@router.post("/import", response_class=HTMLResponse)
async def import_submit(
    request: Request,
    archive: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    filename = archive.filename or ""
    if not filename.lower().endswith(".zip"):
        return templates.TemplateResponse(
            "import.html",
            {
                "request": request,
                "user": user,
                "error": "Ожидается .zip‑архив с .md файлами.",
                "summary": None,
            },
            status_code=400,
        )

    try:
        data = await archive.read()
        zf = zipfile.ZipFile(io.BytesIO(data))
    except Exception:
        return templates.TemplateResponse(
            "import.html",
            {
                "request": request,
                "user": user,
                "error": "Не удалось прочитать архив. Проверь, что это корректный .zip.",
                "summary": None,
            },
            status_code=400,
        )

    imported_count = 0
    skipped_count = 0
    created_questions: List[Question] = []

    for name in zf.namelist():
        if name.endswith("/") or not name.lower().endswith(".md"):
            continue

        try:
            raw_bytes = zf.read(name)
        except KeyError:
            continue

        try:
            raw_text = raw_bytes.decode("utf-8")
        except UnicodeDecodeError:
            raw_text = raw_bytes.decode("cp1251", errors="ignore")

        parsed = parse_markdown_to_question(raw_text)
        if not parsed:
            skipped_count += 1
            continue

        q = Question(
            text=parsed["text"],
            answer_type=parsed["answer_type"],
            options=parsed["options_json"],
            correct=parsed["correct"],
            category=parsed.get("category"),
            grade=parsed.get("grade"),
            year=parsed.get("year"),
            stage=parsed.get("stage"),
        )

        db.add(q)
        db.flush()
        created_questions.append(q)
        imported_count += 1

    db.commit()

    summary = {
        "filename": filename,
        "imported_count": imported_count,
        "skipped_count": skipped_count,
        "created_questions": created_questions,
    }

    return templates.TemplateResponse(
        "import.html",
        {
            "request": request,
            "user": user,
            "error": None,
            "summary": summary,
        },
    )


# ---------- QUESTIONS: список / новая / редактор / удаление ----------


@router.get("/questions", response_class=HTMLResponse)
async def questions_list(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    # забираем все вопросы
    rows: List[Question] = db.query(Question).all()

    # строим иерархию: Категория -> Класс -> Год -> Этап -> список вопросов
    library: dict[str, dict[int, dict[str, dict[str, list[Question]]]]] = {}

    for q in rows:
        category = _category_label(q)
        try:
            grade = int(q.grade) if q.grade is not None else 0
        except (TypeError, ValueError):
            grade = 0
        year = q.year or "?"
        stage = q.stage or "?"

        library.setdefault(category, {}).setdefault(grade, {}).setdefault(year, {}).setdefault(stage, []).append(q)
    return templates.TemplateResponse(
        "questions_list.html",
        {
            "request": request,
            "user": user,
            "library": library,
            "error": None,
            "success": None,
        },
    )



@router.get("/questions/new", response_class=HTMLResponse)
async def question_new_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    categories = _fetch_categories(db)
    years = [
        y[0]
        for y in db.query(Question.year)
        .filter(Question.year.isnot(None), Question.year != "")
        .distinct()
        .order_by(Question.year.asc())
        .all()
    ]
    return templates.TemplateResponse(
        "question_new.html",
        {
            "request": request,
            "user": user,
            "error": None,
            "success": None,
            "categories": _build_category_choices(categories, roots_only=True),
            "new_category_name": "",
            "selected_category_id": None,
            "grade_choices": GRADE_CHOICES,
            "stage_choices": STAGE_CHOICES,
            "year_choices": years,
            "grade": None,
            "year": None,
            "stage": None,
        },
    )


@router.post("/questions/new", response_class=HTMLResponse)
async def question_new_submit(
    request: Request,
    text: str = Form(...),
    answer_type: str = Form(...),
    correct_text: str = Form(""),
    correct_index: str = Form(""),
    correct_number: str = Form(""),
    category_id: Optional[str] = Form(None),
    new_category_name: str = Form(""),
    grade: Optional[str] = Form(None),
    year: Optional[str] = Form(None),
    stage: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    form = await request.form()
    error: Optional[str] = None
    answer_type = answer_type.strip()
    new_category_name = (new_category_name or "").strip()
    selected_category_id_raw = category_id
    grade_val = (grade or "").strip()
    year_val = (year or "").strip()
    stage_val = (stage or "").strip()

    categories_list = _fetch_categories(db)
    category_choices = _build_category_choices(categories_list, roots_only=True)
    years = [
        y[0]
        for y in db.query(Question.year)
        .filter(Question.year.isnot(None), Question.year != "")
        .distinct()
        .order_by(Question.year.asc())
        .all()
    ]

    def collect_options(form_data):
        entries = []
        for k in form_data.keys():
            m = re.match(r"option_(\d+)", k)
            if m:
                idx = int(m.group(1))
                entries.append((idx, (form_data.get(k) or "").strip()))
        entries.sort(key=lambda x: x[0])
        opts = [v for _, v in entries]
        while opts and not opts[-1]:
            opts.pop()
        return opts

    options = collect_options(form)
    pairs: list[dict] = []

    if answer_type not in ("text", "single", "multi", "number", "match"):
        error = "Неподдерживаемый тип ответа."
    elif answer_type == "text":
        if not correct_text.strip():
            error = "Укажите правильный текстовый ответ."
    elif answer_type == "match":
        # пары A-B
        for k in form.keys():
            m = re.match(r"match_left_(\d+)", k)
            if m:
                idx = int(m.group(1))
                left = (form.get(k) or "").strip()
                right = (form.get(f"match_right_{idx}") or "").strip()
                if left and right:
                    pairs.append({"left": left, "right": right})
        if not pairs:
            error = "Добавьте хотя бы одну пару для соотношения."
    elif answer_type in ("single", "multi"):
        # валидные индексы — только для реально заполненных опций
        valid_indices = [i for i, v in enumerate(options) if v.strip()]
        if not valid_indices:
            error = "Добавьте хотя бы один вариант ответа."
        else:
            selected: set[int] = set()
            # чекбоксы
            try:
                selected.update(int(x) for x in form.getlist("correct_multi"))
            except Exception:
                pass
            # радиокнопка
            if correct_index not in ("", None):
                try:
                    selected.add(int(correct_index))
                except ValueError:
                    pass
            # фильтрация по существующим и непустым вариантам
            selected = {i for i in selected if 0 <= i < len(options) and options[i].strip()}
            if not selected:
                # подстрахуемся: берём первый непустой вариант
                selected = {valid_indices[0]}
            normalized = sorted(selected)
            # решаем финальный тип: 1 вариант => single, иначе multi
            answer_type = "single" if len(normalized) == 1 else "multi"
            if answer_type == "single":
                correct_index = str(normalized[0])
            else:
                correct_multi = normalized
    elif answer_type == "number":
        if not correct_number.strip():
            error = "Укажите правильное число."
        else:
            try:
                float(correct_number.strip())
            except ValueError:
                error = "Числовой ответ должен быть числом."

    if error is None:
        if grade_val and grade_val not in GRADE_CHOICES:
            error = "Класс должен быть 9, 10 или 11."
        if stage_val and stage_val not in STAGE_CHOICES:
            error = "Этап должен быть выбран из списка."

    if error:
        return templates.TemplateResponse(
            "question_new.html",
            {
                "request": request,
                "user": user,
                "error": error,
                "success": None,
                "categories": category_choices,
                "selected_category_id": selected_category_id_raw,
                "new_category_name": new_category_name,
                "grade_choices": GRADE_CHOICES,
                "stage_choices": STAGE_CHOICES,
                "year_choices": years,
                "grade": grade_val,
                "year": year_val,
                "stage": stage_val,
            },
            status_code=400,
        )

    options_json = None
    correct = None

    if answer_type == "text":
        correct = correct_text.strip()
    elif answer_type == "number":
        correct = correct_number.strip()
    elif answer_type == "match":
        options_json = json.dumps(pairs, ensure_ascii=False) if pairs else None
        # по умолчанию соотношение 1:1 по строкам
        correct = json.dumps(list(range(len(pairs))), ensure_ascii=False) if pairs else "[]"
    elif answer_type == "multi":
        options_json = json.dumps(options, ensure_ascii=False) if options else None
        correct = json.dumps(sorted(set(correct_multi)), ensure_ascii=False)
    elif answer_type == "single":
        options_json = json.dumps(options, ensure_ascii=False) if options else None
        correct = str(correct_index)

    category_obj: Optional[Category] = None
    if new_category_name:
        category_obj = Category(name=new_category_name, parent_id=None)
        db.add(category_obj)
        try:
            db.flush()
        except IntegrityError:
            db.rollback()
            categories_list = _fetch_categories(db)
            return templates.TemplateResponse(
                "question_new.html",
                {
                    "request": request,
                    "user": user,
                    "error": "Категория с таким именем уже существует на этом уровне.",
                    "success": None,
                    "categories": _build_category_choices(categories_list, roots_only=True),
                    "selected_category_id": selected_category_id_raw,
                    "new_category_name": new_category_name,
                    "grade_choices": GRADE_CHOICES,
                    "stage_choices": STAGE_CHOICES,
                    "year_choices": years,
                    "grade": grade_val,
                    "year": year_val,
                    "stage": stage_val,
                },
                status_code=400,
            )
    elif selected_category_id_raw not in (None, "", "0"):
        try:
            category_obj = db.get(Category, int(selected_category_id_raw))
        except ValueError:
            category_obj = None

    category_label = None
    if category_obj:
        try:
            category_label = category_obj.full_path
        except Exception:
            category_label = category_obj.name

    q = Question(
        text=text,
        answer_type=answer_type,
        options=options_json,
        correct=correct,
        category_id=category_obj.id if category_obj else None,
        category=category_label,
        grade=grade_val or None,
        year=year_val or None,
        stage=stage_val or None,
    )
    db.add(q)
    db.commit()

    return templates.TemplateResponse(
        "question_new.html",
        {
            "request": request,
            "user": user,
            "error": None,
            "success": f"Вопрос успешно сохранён. ID: {q.id}",
            "categories": _build_category_choices(_fetch_categories(db), roots_only=True),
            "selected_category_id": None,
            "new_category_name": "",
            "grade_choices": GRADE_CHOICES,
            "stage_choices": STAGE_CHOICES,
            "year_choices": years,
            "grade": None,
            "year": None,
            "stage": None,
        },
    )

@router.get("/questions/{question_id}/edit", response_class=HTMLResponse)
async def question_edit(
    request: Request,
    question_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    q = db.get(Question, question_id)
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")

    options = []
    correct_multi: list[int] = []
    correct_number = None
    correct_index = None  # поддержка старого поля для контекста шаблона

    match_pairs: list[dict] = []

    if q.answer_type == "match":
        try:
            raw_pairs = json.loads(q.options or "[]")
            for item in raw_pairs:
                left = (item.get("left") if isinstance(item, dict) else None) or ""
                right = (item.get("right") if isinstance(item, dict) else None) or ""
                if left or right:
                    match_pairs.append({"left": left, "right": right})
        except Exception:
            match_pairs = []
        while len(match_pairs) < 4:
            match_pairs.append({"left": "", "right": ""})
    else:
        if q.options:
            try:
                options = json.loads(q.options)
            except Exception:
                options = []
        if q.answer_type == "single" and q.correct is not None:
            try:
                correct_multi = [int(str(q.correct))]
            except ValueError:
                correct_multi = []
        if q.answer_type == "multi" and q.correct:
            try:
                correct_multi = json.loads(q.correct) if q.correct else []
            except Exception:
                correct_multi = []
        if q.answer_type == "number":
            correct_number = q.correct

        while len(options) < 4:
            options.append("")

    years = [
        y[0]
        for y in db.query(Question.year)
        .filter(Question.year.isnot(None), Question.year != "")
        .distinct()
        .order_by(Question.year.asc())
        .all()
    ]
    categories = _fetch_categories(db)
    return templates.TemplateResponse(
        "question_edit.html",
        {
            "request": request,
            "user": user,
            "question": q,
            "options": options,
            "correct_index": correct_index,
            "correct_multi": correct_multi,
            "correct_number": correct_number,
            "match_pairs": match_pairs,
            "error": None,
            "success": None,
            "categories": _build_category_choices(categories, roots_only=True),
            "selected_category_id": getattr(q, "category_id", None),
            "new_category_name": "",
            "category_label": _category_label(q),
            "grade_choices": GRADE_CHOICES,
            "stage_choices": STAGE_CHOICES,
            "year_choices": years,
            "grade": getattr(q, "grade", None),
            "year": getattr(q, "year", None),
            "stage": getattr(q, "stage", None),
        },
    )


@router.post("/questions/{question_id}/edit", response_class=HTMLResponse)
async def question_edit_post(
    request: Request,
    question_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    q = db.get(Question, question_id)
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")

    form = await request.form()
    text = (form.get("text") or "").strip()
    answer_type = (form.get("answer_type") or "text").strip()
    new_category_name = (form.get("new_category_name") or "").strip()
    selected_category_id_raw = form.get("category_id")
    grade_val = (form.get("grade") or "").strip()
    year_val = (form.get("year") or "").strip()
    stage_val = (form.get("stage") or "").strip()
    categories_list = _fetch_categories(db)
    category_choices = _build_category_choices(categories_list, roots_only=True)
    years = [
        y[0]
        for y in db.query(Question.year)
        .filter(Question.year.isnot(None), Question.year != "")
        .distinct()
        .order_by(Question.year.asc())
        .all()
    ]

    def collect_options(form_data):
        entries = []
        for k in form_data.keys():
            m = re.match(r"option_(\d+)", k)
            if m:
                idx = int(m.group(1))
                entries.append((idx, (form_data.get(k) or "").strip()))
        entries.sort(key=lambda x: x[0])
        opts = [v for _, v in entries]
        while opts and not opts[-1]:
            opts.pop()
        return opts

    if not text:
        return templates.TemplateResponse(
            "question_edit.html",
            {
                "request": request,
                "user": user,
                "question": q,
                "options": collect_options(form),
                "correct_index": None,
                "correct_multi": [],
                "correct_number": None,
                "error": "Текст вопроса обязателен",
                "success": None,
                "match_pairs": [{"left": "", "right": ""} for _ in range(4)],
                "categories": category_choices,
                "selected_category_id": selected_category_id_raw or getattr(q, "category_id", None),
                "new_category_name": new_category_name,
                "category_label": _category_label(q),
                "grade_choices": GRADE_CHOICES,
                "stage_choices": STAGE_CHOICES,
                "year_choices": years,
                "grade": grade_val or getattr(q, "grade", None),
                "year": year_val or getattr(q, "year", None),
                "stage": stage_val or getattr(q, "stage", None),
            },
        )

    q.text = text
    q.answer_type = answer_type

    if answer_type == "text":
        q.correct = (form.get("correct_text") or "").strip()
        q.options = None
    elif answer_type == "number":
        num_raw = (form.get("correct_number") or "").strip()
        if not num_raw:
            return templates.TemplateResponse(
                "question_edit.html",
                {
                    "request": request,
                    "user": user,
                    "question": q,
                    "options": collect_options(form),
                    "correct_index": None,
                    "correct_multi": [],
                    "correct_number": None,
                    "match_pairs": [],
                    "error": "Укажите правильное число",
                    "success": None,
                    "categories": category_choices,
                    "selected_category_id": selected_category_id_raw or getattr(q, "category_id", None),
                    "new_category_name": new_category_name,
                    "category_label": _category_label(q),
                    "grade_choices": GRADE_CHOICES,
                    "stage_choices": STAGE_CHOICES,
                    "year_choices": years,
                    "grade": grade_val or getattr(q, "grade", None),
                    "year": year_val or getattr(q, "year", None),
                    "stage": stage_val or getattr(q, "stage", None),
                },
                status_code=400,
            )
        try:
            float(num_raw)
        except ValueError:
            return templates.TemplateResponse(
                "question_edit.html",
                {
                    "request": request,
                    "user": user,
                    "question": q,
                    "options": collect_options(form),
                    "correct_index": None,
                    "correct_multi": [],
                    "correct_number": None,
                    "match_pairs": [],
                    "error": "Числовой ответ должен быть числом",
                    "success": None,
                    "categories": category_choices,
                    "selected_category_id": selected_category_id_raw or getattr(q, "category_id", None),
                    "new_category_name": new_category_name,
                    "category_label": _category_label(q),
                    "grade_choices": GRADE_CHOICES,
                    "stage_choices": STAGE_CHOICES,
                    "year_choices": years,
                    "grade": grade_val or getattr(q, "grade", None),
                    "year": year_val or getattr(q, "year", None),
                    "stage": stage_val or getattr(q, "stage", None),
                },
                status_code=400,
            )
        q.correct = num_raw
        q.options = None
    elif answer_type == "match":
        pairs = []
        for k in form.keys():
            m = re.match(r"match_left_(\d+)", k)
            if m:
                idx = int(m.group(1))
                left = (form.get(k) or "").strip()
                right = (form.get(f"match_right_{idx}") or "").strip()
                if left and right:
                    pairs.append({"left": left, "right": right})
        if not pairs:
            return templates.TemplateResponse(
                "question_edit.html",
                {
                    "request": request,
                    "user": user,
                    "question": q,
                    "options": collect_options(form),
                    "correct_index": None,
                    "correct_multi": [],
                    "correct_number": None,
                    "match_pairs": [{"left": "", "right": ""} for _ in range(4)],
                    "error": "Добавьте пары для соотношения",
                    "success": None,
                    "categories": category_choices,
                    "selected_category_id": selected_category_id_raw or getattr(q, "category_id", None),
                    "new_category_name": new_category_name,
                    "category_label": _category_label(q),
                    "grade_choices": GRADE_CHOICES,
                    "stage_choices": STAGE_CHOICES,
                    "year_choices": years,
                    "grade": grade_val or getattr(q, "grade", None),
                    "year": year_val or getattr(q, "year", None),
                    "stage": stage_val or getattr(q, "stage", None),
                },
                status_code=400,
            )
        q.answer_type = "match"
        q.options = json.dumps(pairs, ensure_ascii=False)
        q.correct = json.dumps(list(range(len(pairs))), ensure_ascii=False)
    elif answer_type in ("single", "multi"):
        options = collect_options(form)
        valid_indices = [i for i, v in enumerate(options) if str(v).strip()]
        if not valid_indices:
            return templates.TemplateResponse(
                "question_edit.html",
                {
                    "request": request,
                    "user": user,
                    "question": q,
                    "options": options,
                    "correct_index": None,
                    "correct_multi": [],
                    "correct_number": None,
                    "error": "Добавьте хотя бы один вариант ответа",
                    "success": None,
                    "match_pairs": [{"left": "", "right": ""} for _ in range(4)],
                    "categories": category_choices,
                    "selected_category_id": selected_category_id_raw or getattr(q, "category_id", None),
                    "new_category_name": new_category_name,
                    "category_label": _category_label(q),
                    "grade_choices": GRADE_CHOICES,
                    "stage_choices": STAGE_CHOICES,
                    "year_choices": years,
                    "grade": grade_val or getattr(q, "grade", None),
                    "year": year_val or getattr(q, "year", None),
                    "stage": stage_val or getattr(q, "stage", None),
                },
                status_code=400,
            )
        q.options = json.dumps(options, ensure_ascii=False) if options else None

        selected: set[int] = set()
        try:
            selected.update(int(x) for x in form.getlist("correct_multi"))
        except Exception:
            pass
        correct_raw = form.get("correct_index")
        if correct_raw not in ("", None):
            try:
                selected.add(int(correct_raw))
            except ValueError:
                pass

        selected = {i for i in selected if 0 <= i < len(options) and options[i].strip()}
        if not selected:
            selected = {valid_indices[0]}

        normalized = sorted(selected)
        answer_type = "single" if len(normalized) == 1 else "multi"
        q.answer_type = answer_type
        if answer_type == "single":
            q.correct = str(normalized[0])
        else:
            q.correct = json.dumps(normalized, ensure_ascii=False)
    else:
        return templates.TemplateResponse(
            "question_edit.html",
            {
                "request": request,
                "user": user,
                "question": q,
                "options": [],
                "correct_index": None,
                "correct_multi": [],
                "correct_number": None,
                "error": "Неподдерживаемый тип ответа",
                "success": None,
                "match_pairs": [{"left": "", "right": ""} for _ in range(4)],
                "categories": category_choices,
                "selected_category_id": selected_category_id_raw or getattr(q, "category_id", None),
                "new_category_name": new_category_name,
                "category_label": _category_label(q),
                "grade_choices": GRADE_CHOICES,
                "stage_choices": STAGE_CHOICES,
                "year_choices": years,
                "grade": grade_val or getattr(q, "grade", None),
                "year": year_val or getattr(q, "year", None),
                "stage": stage_val or getattr(q, "stage", None),
            },
            status_code=400,
        )

    if grade_val and grade_val not in GRADE_CHOICES:
        return templates.TemplateResponse(
            "question_edit.html",
            {
                "request": request,
                "user": user,
                "question": q,
                "options": collect_options(form),
                "correct_index": None,
                "correct_multi": [],
                "correct_number": None,
                "match_pairs": [{"left": "", "right": ""} for _ in range(4)],
                "error": "Класс должен быть 9, 10 или 11.",
                "success": None,
                "categories": category_choices,
                "selected_category_id": selected_category_id_raw or getattr(q, "category_id", None),
                "new_category_name": new_category_name,
                "category_label": _category_label(q),
                "grade_choices": GRADE_CHOICES,
                "stage_choices": STAGE_CHOICES,
                "year_choices": years,
                "grade": grade_val or getattr(q, "grade", None),
                "year": year_val or getattr(q, "year", None),
                "stage": stage_val or getattr(q, "stage", None),
            },
            status_code=400,
        )
    if stage_val and stage_val not in STAGE_CHOICES:
        return templates.TemplateResponse(
            "question_edit.html",
            {
                "request": request,
                "user": user,
                "question": q,
                "options": collect_options(form),
                "correct_index": None,
                "correct_multi": [],
                "correct_number": None,
                "match_pairs": [{"left": "", "right": ""} for _ in range(4)],
                "error": "Этап должен быть выбран из списка.",
                "success": None,
                "categories": category_choices,
                "selected_category_id": selected_category_id_raw or getattr(q, "category_id", None),
                "new_category_name": new_category_name,
                "category_label": _category_label(q),
                "grade_choices": GRADE_CHOICES,
                "stage_choices": STAGE_CHOICES,
                "year_choices": years,
                "grade": grade_val or getattr(q, "grade", None),
                "year": year_val or getattr(q, "year", None),
                "stage": stage_val or getattr(q, "stage", None),
            },
            status_code=400,
        )

    category_obj: Optional[Category] = None
    if new_category_name:
        category_obj = Category(name=new_category_name, parent_id=None)
        db.add(category_obj)
        try:
            db.flush()
        except IntegrityError:
            db.rollback()
            categories_list = _fetch_categories(db)
            return templates.TemplateResponse(
                "question_edit.html",
                {
                    "request": request,
                    "user": user,
                    "question": q,
                    "options": collect_options(form),
                    "correct_index": None,
                    "correct_multi": [],
                    "correct_number": None,
                    "match_pairs": [{"left": "", "right": ""} for _ in range(4)],
                    "error": "Категория с таким именем уже есть на этом уровне.",
                    "success": None,
                    "categories": _build_category_choices(categories_list, roots_only=True),
                    "selected_category_id": selected_category_id_raw or getattr(q, "category_id", None),
                    "new_category_name": new_category_name,
                    "category_label": _category_label(q),
                    "grade_choices": GRADE_CHOICES,
                    "stage_choices": STAGE_CHOICES,
                    "year_choices": years,
                    "grade": grade_val or getattr(q, "grade", None),
                    "year": year_val or getattr(q, "year", None),
                    "stage": stage_val or getattr(q, "stage", None),
                },
                status_code=400,
            )
    elif selected_category_id_raw not in (None, "", "0"):
        try:
            category_obj = db.get(Category, int(selected_category_id_raw))
        except ValueError:
            category_obj = None
    else:
        category_obj = db.get(Category, getattr(q, "category_id", None)) if getattr(q, "category_id", None) else None

    if category_obj:
        try:
            q.category = category_obj.full_path
        except Exception:
            q.category = category_obj.name
        q.category_id = category_obj.id
    else:
        q.category_id = None
        q.category = None

    q.grade = grade_val or None
    q.year = year_val or None
    q.stage = stage_val or None

    db.add(q)
    db.commit()
    db.refresh(q)

    options = []
    correct_index = None
    correct_multi = []
    correct_number = None
    match_pairs: list[dict] = []
    if q.answer_type == "match":
        try:
            for item in json.loads(q.options or "[]"):
                left = (item.get("left") if isinstance(item, dict) else None) or ""
                right = (item.get("right") if isinstance(item, dict) else None) or ""
                match_pairs.append({"left": left, "right": right})
        except Exception:
            match_pairs = []
    else:
        if q.options:
            try:
                options = json.loads(q.options)
            except Exception:
                options = []
        if q.answer_type == "single" and q.correct is not None:
            try:
                correct_index = int(str(q.correct))
            except ValueError:
                correct_index = None
        if q.answer_type == "multi" and q.correct:
            try:
                correct_multi = json.loads(q.correct) if q.correct else []
            except Exception:
                correct_multi = []
        if q.answer_type == "number":
            correct_number = q.correct
    while len(options) < 4:
        options.append("")
    while len(match_pairs) < 4:
        match_pairs.append({"left": "", "right": ""})

    return templates.TemplateResponse(
        "question_edit.html",
        {
            "request": request,
            "user": user,
            "question": q,
            "options": options,
            "correct_index": correct_index,
            "correct_multi": correct_multi,
            "correct_number": correct_number,
            "match_pairs": match_pairs,
            "error": None,
            "success": "Изменения сохранены",
            "categories": _build_category_choices(_fetch_categories(db), roots_only=True),
            "selected_category_id": getattr(q, "category_id", None),
            "new_category_name": "",
            "category_label": _category_label(q),
            "grade_choices": GRADE_CHOICES,
            "stage_choices": STAGE_CHOICES,
            "year_choices": years,
            "grade": getattr(q, "grade", None),
            "year": getattr(q, "year", None),
            "stage": getattr(q, "stage", None),
        },
    )

@router.post("/questions/{question_id}/delete", response_class=HTMLResponse)
async def question_delete(
    question_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    q = db.get(Question, question_id)
    if not q:
        error = f"Вопрос #{question_id} не найден."
        rows: List[Question] = db.query(Question).order_by(Question.id.desc()).all()
        return templates.TemplateResponse(
            "questions_list.html",
            {
                "request": request,
                "user": user,
                "questions": rows,
                "error": error,
                "success": None,
            },
            status_code=400,
        )

    db.query(Answer).filter(Answer.question_id == question_id).delete()
    db.query(TestQuestion).filter(TestQuestion.question_id == question_id).delete()
    db.delete(q)
    db.commit()

    rows: List[Question] = db.query(Question).order_by(Question.id.desc()).all()
    return templates.TemplateResponse(
        "questions_list.html",
        {
            "request": request,
            "user": user,
            "questions": rows,
            "error": None,
            "success": f"Вопрос #{question_id} удалён.",
        },
    )


# ---------- TESTS UI ----------


@router.get("/tests", response_class=HTMLResponse)
async def tests_list(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    role = user.role if user else None
    query = db.query(Test)
    if role not in ("teacher", "admin"):
        query = query.filter(
            or_(
                Test.description.is_(None),
                not_(Test.description.ilike("Автосгенерированный%")),
            )
        )
    tests = query.order_by(Test.id.desc()).all()

    def build_info(test_list: List[Test]) -> List[dict]:
        out = []
        for t in test_list:
            tqs: List[TestQuestion] = (
                db.query(TestQuestion)
                .filter(TestQuestion.test_id == t.id)
                .all()
            )
            out.append(
                {
                    "test": t,
                    "question_count": len(tqs),
                    "max_score": sum(tq.points for tq in tqs) if tqs else 0,
                }
            )
        return out

    submissions = (
        db.query(Submission)
        .filter(Submission.user_id == user.id)
        .all()
        if user
        else []
    )
    passed_ids = {s.test_id for s in submissions if s.test_id}
    passed_tests = [t for t in tests if t.id in passed_ids]

    if role in ("teacher", "admin"):
        new_tests = tests
    else:
        new_tests = [t for t in tests if (t.id not in passed_ids) and ((t.created_by and t.created_by.role == "admin") or t.created_by_id is None)]

    context = {
        "request": request,
        "user": user,
        "role": role,
        "new_tests": build_info(new_tests),
        "passed_tests": build_info(passed_tests),
    }
    return templates.TemplateResponse(
        "tests_list.html",
        context,
    )


def _collect_test_stats(db: Session, test: Test) -> dict:
    tqs: List[TestQuestion] = (
        db.query(TestQuestion)
        .filter(TestQuestion.test_id == test.id)
        .order_by(TestQuestion.order.asc(), TestQuestion.id.asc())
        .all()
    )
    max_score = sum(tq.points for tq in tqs) if tqs else 0

    submissions: List[Submission] = (
        db.query(Submission)
        .join(User, Submission.user_id == User.id)
        .filter(Submission.test_id == test.id, User.role == UserRole.STUDENT)
        .order_by(Submission.id.desc())
        .all()
    )
    latest_by_user: dict[int, Submission] = {}
    for sub in submissions:
        if sub.user_id not in latest_by_user:
            latest_by_user[sub.user_id] = sub
    submissions = list(latest_by_user.values())

    sub_ids = [s.id for s in submissions]
    answers_by_sub: dict[int, dict[int, Answer]] = {}
    answers_by_question: dict[int, list[Answer]] = {}
    if sub_ids:
        answers = db.query(Answer).filter(Answer.submission_id.in_(sub_ids)).all()
        for a in answers:
            answers_by_sub.setdefault(a.submission_id, {})[a.question_id] = a
            answers_by_question.setdefault(a.question_id, []).append(a)

    user_ids = [s.user_id for s in submissions]
    users_map: dict[int, User] = (
        {u.id: u for u in db.query(User).filter(User.id.in_(user_ids)).all()}
        if user_ids
        else {}
    )

    question_stats = []
    overall_correct = 0
    overall_total = 0
    for idx, tq in enumerate(tqs, 1):
        q = getattr(tq, "question", None) or db.get(Question, tq.question_id)
        text = (getattr(q, "text", "") or "").replace("\n", " ").strip()
        if len(text) > 120:
            text = text[:117] + "..."
        ans_list = answers_by_question.get(tq.question_id, [])
        total = len(ans_list)
        correct = sum(1 for a in ans_list if getattr(a, "correct", False))
        overall_total += total
        overall_correct += correct
        question_stats.append(
            {
                "order": idx,
                "question": q,
                "question_text": text,
                "correct": correct,
                "wrong": total - correct,
                "total": total,
                "points": tq.points,
                "dots": [bool(getattr(a, "correct", False)) for a in ans_list],
            }
        )

    student_rows = []
    for sub in submissions:
        u = users_map.get(sub.user_id)
        if not u:
            continue
        ans_map = answers_by_sub.get(sub.id, {})
        flags: list[bool] = []
        for tq in tqs:
            a = ans_map.get(tq.question_id)
            flags.append(bool(a.correct) if a else False)
        percent = None
        if max_score:
            percent = round(((sub.score or 0) / max_score) * 100, 1)
        student_rows.append(
            {
                "user": u,
                "submission": sub,
                "score": sub.score or 0,
                "max_score": max_score,
                "percent": percent,
                "answers": flags,
            }
        )

    student_rows.sort(
        key=lambda r: (
            -(r["score"] or 0),
            (getattr(r["user"], "full_name", "") or r["user"].email).lower(),
        )
    )

    overall_wrong = max(overall_total - overall_correct, 0)

    return {
        "questions": tqs,
        "max_score": max_score,
        "submissions": submissions,
        "users_map": users_map,
        "question_stats": question_stats,
        "student_rows": student_rows,
        "overall": {
            "total_answers": overall_total,
            "correct": overall_correct,
            "wrong": overall_wrong,
            "accuracy": round((overall_correct / overall_total) * 100, 1)
            if overall_total
            else None,
        },
    }


@router.get("/tests/{test_id}/stats", response_class=HTMLResponse)
async def test_stats_view(
    test_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    test = db.get(Test, test_id)
    if not test:
        raise HTTPException(status_code=404, detail="test not found")
    stats = _collect_test_stats(db, test)
    return templates.TemplateResponse(
        "test_stats.html",
        {
            "request": request,
            "user": user,
            "test": test,
            **stats,
        },
    )


@router.get("/tests/{test_id}/stats/export")
async def test_stats_export(
    test_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    test = db.get(Test, test_id)
    if not test:
        raise HTTPException(status_code=404, detail="test not found")
    stats = _collect_test_stats(db, test)
    output = io.StringIO()
    writer = csv.writer(output)
    header = ["ФИО", "Почта", "Класс", "Баллы", "Максимум", "Процент"]
    for qs in stats["question_stats"]:
        header.append(f"Вопрос {qs['order']}")
    writer.writerow(header)

    for row in stats["student_rows"]:
        u: User = row["user"]
        percent = row["percent"] if row["percent"] is not None else ""
        line = [
            getattr(u, "full_name", "") or u.email,
            u.email,
            getattr(u, "student_class", "") or "",
            row["score"],
            row["max_score"],
            percent,
        ]
        for flag in row["answers"]:
            line.append(1 if flag else 0)
        writer.writerow(line)

    csv_bytes = output.getvalue().encode("utf-8-sig")
    filename = f"test_{test.id}_stats.csv"
    return Response(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.get("/tests/random", response_class=HTMLResponse)
async def random_test_form(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    categories = _get_root_category_choices(db)
    return templates.TemplateResponse(
        "random_test.html",
        {
            "request": request,
            "user": user,
            "categories": categories,
            "grade_choices": GRADE_CHOICES,
            "selected_categories": [],
            "grade": None,
            "count": 10,
            "error": None,
        },
    )


@router.post("/tests/random", response_class=HTMLResponse)
async def random_test_submit(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    form = await request.form()
    selected_categories: list[int] = []
    for v in form.getlist("category_ids"):
        try:
            selected_categories.append(int(v))
        except Exception:
            continue
    grade_val = (form.get("grade") or "").strip()
    try:
        count = int(form.get("count") or 10)
    except ValueError:
        count = 10
    count = max(1, min(count, 50))

    if grade_val and grade_val not in GRADE_CHOICES:
        grade_val = ""

    query = db.query(Question)
    if selected_categories:
        query = query.filter(Question.category_id.in_(selected_categories))
    if grade_val:
        query = query.filter(Question.grade == grade_val)

    questions = query.order_by(func.random()).limit(count).all()
    if not questions:
        categories = _get_root_category_choices(db)
        return templates.TemplateResponse(
            "random_test.html",
            {
                "request": request,
                "user": user,
                "categories": categories,
                "grade_choices": GRADE_CHOICES,
                "error": "Не найдено ни одного вопроса по выбранным условиям.",
                "selected_categories": selected_categories,
                "grade": grade_val,
                "count": count,
            },
            status_code=400,
        )

    title_parts = ["Случайный тест"]
    if grade_val:
        title_parts.append(f"{grade_val} класс")
    test = Test(
        title=" / ".join(title_parts),
        description="Автосгенерированный тест по выбранным категориям",
        created_by_id=user.id if hasattr(user, "id") else None,
        is_public=False,
        show_answers_to_student=True,
    )
    db.add(test)
    db.flush()

    order = 0
    for q in questions:
        order += 1
        db.add(
            TestQuestion(
                test_id=test.id,
                question_id=q.id,
                points=1,
                order=order,
            )
        )
    db.commit()

    return redirect(f"/ui/tests/run/{test.id}")


@router.get("/tests/new", response_class=HTMLResponse)
async def test_builder_new(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    questions: List[Question] = db.query(Question).order_by(Question.id.asc()).all()
    library: dict[str, dict[int, dict[str, dict[str, list[Question]]]]] = {}
    for q in questions:
        category = _category_label(q)
        try:
            grade = int(q.grade) if q.grade is not None else 0
        except (TypeError, ValueError):
            grade = 0
        year = q.year or ""
        stage = q.stage or ""
        library.setdefault(category, {}).setdefault(grade, {}).setdefault(year, {}).setdefault(stage, []).append(q)
    return templates.TemplateResponse(
        "test_builder.html",
        {
            "request": request,
            "user": user,
            "mode": "create",
            "test": None,
            "questions": questions,
            "library": library,
            "selected": {},
        },
    )


@router.post("/tests/new", response_class=HTMLResponse)
async def test_builder_new_post(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    form = await request.form()
    title = (form.get("title") or "").strip()
    description = (form.get("description") or "").strip()
    show_correct = "show_correct_answers" in form
    try:
        max_attempts = int(form.get("max_attempts") or 0)
    except ValueError:
        max_attempts = 0
    if max_attempts < 0:
        max_attempts = 0

    questions: List[Question] = db.query(Question).order_by(Question.id.asc()).all()
    selection: list[tuple[int, int]] = []
    for q in questions:
        if form.get(f"q_{q.id}_include") is None:
            continue
        try:
            points = int(form.get(f"q_{q.id}_points") or 1)
        except ValueError:
            points = 1
        selection.append((q.id, max(points, 0)))

    if not title:
        error = "Укажите название теста"
    elif not selection:
        error = "Выберите хотя бы один вопрос"
    else:
        error = None

    if error:
        library: dict[str, dict[int, dict[str, dict[str, list[Question]]]]] = {}
        for q in questions:
            category = _category_label(q)
            try:
                grade = int(q.grade) if q.grade is not None else 0
            except (TypeError, ValueError):
                grade = 0
            year = q.year or ""
            stage = q.stage or ""
            library.setdefault(category, {}).setdefault(grade, {}).setdefault(year, {}).setdefault(stage, []).append(q)
        return templates.TemplateResponse(
            "test_builder.html",
            {
                "request": request,
                "user": user,
                "mode": "create",
                "test": None,
                "questions": questions,
                "library": library,
                "selected": {},
                "max_attempts": max_attempts,
                "error": error,
            },
            status_code=400,
        )

    t = Test(
        title=title,
        description=description or None,
        show_answers_to_student=show_correct,
        created_by_id=user.id if hasattr(user, "id") else None,
        max_attempts=max_attempts or None,
    )
    db.add(t)
    db.flush()

    order = 0
    for question_id, points in selection:
        order += 1
        tq = TestQuestion(
            test_id=t.id,
            question_id=question_id,
            points=points,
            order=order,
        )
        db.add(tq)

    db.commit()
    return redirect("/ui/tests")


@router.get("/tests/{test_id}/edit", response_class=HTMLResponse)
async def test_builder_edit(
    test_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    test = db.get(Test, test_id)
    if not test:
        raise HTTPException(status_code=404, detail="test not found")

    questions: List[Question] = db.query(Question).order_by(Question.id.asc()).all()
    library: dict[str, dict[int, dict[str, dict[str, list[Question]]]]] = {}
    for q in questions:
        category = _category_label(q)
        try:
            grade = int(q.grade) if q.grade is not None else 0
        except (TypeError, ValueError):
            grade = 0
        year = q.year or ""
        stage = q.stage or ""
        library.setdefault(category, {}).setdefault(grade, {}).setdefault(year, {}).setdefault(stage, []).append(q)
    tqs: List[TestQuestion] = (
        db.query(TestQuestion)
        .filter(TestQuestion.test_id == test.id)
        .all()
    )
    selected = {tq.question_id: tq for tq in tqs}

    return templates.TemplateResponse(
        "test_builder.html",
        {
            "request": request,
            "user": user,
            "mode": "edit",
            "test": test,
            "questions": questions,
            "library": library,
            "selected": selected,
        },
    )


@router.post("/tests/{test_id}/edit", response_class=HTMLResponse)
async def test_builder_edit_post(
    test_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    test = db.get(Test, test_id)
    if not test:
        raise HTTPException(status_code=404, detail="test not found")

    form = await request.form()
    title = (form.get("title") or "").strip()
    description = (form.get("description") or "").strip()
    show_correct = "show_correct_answers" in form
    try:
        max_attempts = int(form.get("max_attempts") or 0)
    except ValueError:
        max_attempts = 0
    if max_attempts < 0:
        max_attempts = 0

    questions: List[Question] = db.query(Question).order_by(Question.id.asc()).all()
    selection: list[tuple[int, int]] = []
    for q in questions:
        if form.get(f"q_{q.id}_include") is None:
            continue
        try:
            points = int(form.get(f"q_{q.id}_points") or 1)
        except ValueError:
            points = 1
        selection.append((q.id, max(points, 0)))

    if not title:
        error = "Укажите название теста"
    elif not selection:
        error = "Выберите хотя бы один вопрос"
    else:
        error = None

    if error:
        tqs: List[TestQuestion] = (
            db.query(TestQuestion)
            .filter(TestQuestion.test_id == test.id)
            .all()
        )
        selected = {tq.question_id: tq for tq in tqs}
        library: dict[str, dict[int, dict[str, dict[str, list[Question]]]]] = {}
        for q in questions:
            category = _category_label(q)
            try:
                grade = int(q.grade) if q.grade is not None else 0
            except (TypeError, ValueError):
                grade = 0
            year = q.year or ""
            stage = q.stage or ""
            library.setdefault(category, {}).setdefault(grade, {}).setdefault(year, {}).setdefault(stage, []).append(q)
        return templates.TemplateResponse(
            "test_builder.html",
            {
                "request": request,
                "user": user,
                "mode": "edit",
                "test": test,
                "questions": questions,
                "library": library,
                "selected": selected,
                "max_attempts": max_attempts,
                "error": error,
            },
            status_code=400,
        )

    test.title = title
    test.description = description or None
    test.show_answers_to_student = show_correct
    test.max_attempts = max_attempts or None
    db.add(test)

    db.query(TestQuestion).filter(TestQuestion.test_id == test.id).delete()

    order = 0
    for question_id, points in selection:
        order += 1
        tq = TestQuestion(
            test_id=test.id,
            question_id=question_id,
            points=points,
            order=order,
        )
        db.add(tq)

    db.commit()
    return redirect("/ui/tests")


@router.post("/tests/{test_id}/delete", response_class=HTMLResponse)
async def test_delete(
    test_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    test = db.get(Test, test_id)
    if not test:
        raise HTTPException(status_code=404, detail="test not found")
    # удалить попытки и ответы
    subs: List[Submission] = db.query(Submission).filter(Submission.test_id == test.id).all()
    sub_ids = [s.id for s in subs]
    if sub_ids:
        db.query(Answer).filter(Answer.submission_id.in_(sub_ids)).delete(synchronize_session=False)
        db.query(Submission).filter(Submission.id.in_(sub_ids)).delete(synchronize_session=False)
    db.query(TestQuestion).filter(TestQuestion.test_id == test.id).delete()
    db.delete(test)
    db.commit()
    return redirect("/ui/tests")


@router.get("/tests/{test_id}", response_class=HTMLResponse)
async def test_view(
    test_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    submission_id: Optional[int] = None,
):
    test = db.get(Test, test_id)
    if not test:
        raise HTTPException(status_code=404, detail="test not found")

    tqs: List[TestQuestion] = (
        db.query(TestQuestion)
        .filter(TestQuestion.test_id == test_id)
        .all()
    )
    items = []
    max_points = 0
    for tq in tqs:
        q = db.get(Question, tq.question_id)
        if not q:
            continue
        opts = json.loads(q.options) if q.options else None
        max_points += tq.points
        items.append({"tq": tq, "q": q, "options": opts, "given": None, "earned": 0})

    if not items:
        raise HTTPException(status_code=400, detail="test has no questions")

    # Render first question for the run page
    current = items[0]
    question = current["q"]
    answers_list = []
    if question.answer_type != "text":
        # prefer structured options if present
        if current["options"]:
            answers_list = [
                type("Opt", (), {"id": idx, "text": opt})
                for idx, opt in enumerate(current["options"])
            ]
        elif hasattr(question, "answers") and question.answers:
            answers_list = question.answers

    submission = None
    if submission_id is not None:
        submission = db.get(Submission, submission_id)
        if not submission or submission.user_id != user.id:
            submission = None

    # простая замена markdown изображений на <img>
    def md_to_html(text: str) -> str:
        if not text:
            return ""
        # превращаем ![](url) в <img>, допускаем пробелы вокруг url
        html = re.sub(
            r"!\[[^\]]*\]\(\s*([^)]+?)\s*\)",
            r'<img src="\1" style="max-width:100%;height:auto;">',
            text,
        )
        html = html.replace("\n", "<br>")
        return html

    return templates.TemplateResponse(
        "test_run.html",
        {
            "request": request,
            "user": user,
            "test": test,
            "items": items,
            "question": question,
            "question_html": md_to_html(str(getattr(question, "text", "") or "")),
            "answers": answers_list,
            "answers_html": [md_to_html(str(getattr(a, "text", "") or "")) for a in answers_list] if answers_list else None,
            "index": 0,
            "total_questions": len(items),
            "state_json": "",
            "selected_answer_id": None,
            "selected_answer_ids": [],
            "answer_text": "",
            "max_points": max_points,
            "submission": submission,
            "result": None,
        },
    )


@router.post("/tests/{test_id}/add-question")
async def test_add_question(
    test_id: int,
    question_id: int = Form(...),
    points: int = Form(1),
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    test = db.get(Test, test_id)
    if not test:
        raise HTTPException(status_code=404, detail="test not found")

    q = db.get(Question, question_id)
    if not q:
        raise HTTPException(status_code=404, detail="question not found")

    tq = TestQuestion(test_id=test_id, question_id=question_id, points=points)
    db.add(tq)
    db.commit()
    return redirect(f"/ui/tests/{test_id}")


@router.post("/tests/{test_id}/start")
async def test_start(
    test_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    test = db.get(Test, test_id)
    if not test:
        raise HTTPException(status_code=404, detail="test not found")

    s = Submission(test_id=test_id, user_id=user.id, score=0)
    db.add(s)
    db.commit()
    return redirect(f"/ui/tests/{test_id}?submission_id={s.id}")


@router.get("/tests/{test_id}/start")
async def test_start_get(
    test_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    # Allow starting a test via GET (link click) by delegating to the POST handler.
    return await test_start(test_id=test_id, db=db, user=user)


@router.post("/tests/{test_id}/submit", response_class=HTMLResponse)
async def test_submit(
    test_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    form = await request.form()
    submission_id = int(form.get("submission_id", "0"))
    submission = db.get(Submission, submission_id)

    if not submission or submission.user_id != user.id or submission.test_id != test_id:
        raise HTTPException(status_code=400, detail="invalid submission")

    tqs: List[TestQuestion] = (
        db.query(TestQuestion)
        .filter(TestQuestion.test_id == test_id)
        .all()
    )

    items = []
    max_points = 0
    score = 0

    for tq in tqs:
        q = db.get(Question, tq.question_id)
        if not q:
            continue

        field_name = f"answer_{q.id}"
        opts = json.loads(q.options) if q.options else None
        max_points += tq.points

        given = form.get(field_name, "").strip()
        if not given:
            correct_flag = 0
            earned = 0
        else:
            if q.answer_type == "text":
                gt = (q.correct or "").strip().lower()
                uv = given.strip().lower()
                ok = gt == uv
                correct_flag = 1 if ok else 0
                earned = tq.points if ok else 0
            elif q.answer_type == "single":
                ok = q.correct == given
                correct_flag = 1 if ok else 0
                earned = tq.points if ok else 0
            else:
                correct_flag = 0
                earned = 0

        ans = Answer(
            submission_id=submission.id,
            question_id=q.id,
            given=given,
            correct=bool(correct_flag),
            points=earned,
        )
        db.add(ans)
        score += earned

        items.append({"tq": tq, "q": q, "options": opts})

    submission.score = score
    db.add(submission)
    db.commit()

    test = db.get(Test, test_id)
    result = {"score": score, "max_points": max_points}

    if test.show_answers_to_student:
        rows = []
        for idx, item in enumerate(items, 1):
            q = item["q"]
            opts = item["options"]
            given_raw = item.get("given") or ""

            your_answer = given_raw or "—"
            correct_answer = ""

            if q.answer_type == "text":
                correct_answer = (q.correct or "").strip()
            elif q.answer_type == "single":
                # map numeric index to option text
                try:
                    c_idx = int(q.correct) if q.correct is not None else None
                except (TypeError, ValueError):
                    c_idx = None
                if c_idx is not None and opts and 0 <= c_idx < len(opts):
                    correct_answer = str(opts[c_idx])
                else:
                    correct_answer = str(q.correct or "")

                try:
                    g_idx = int(given_raw)
                    if opts and 0 <= g_idx < len(opts):
                        your_answer = str(opts[g_idx])
                except (TypeError, ValueError):
                    pass

            rows.append(
                {
                    "index": idx,
                    "question": q,
                    "your_answer": your_answer,
                    "correct_answer": correct_answer,
                    "score": item.get("earned", 0),
                    "max_points": next((tq.points for tq in tqs if tq.question_id == q.id), 0),
                }
            )

        return templates.TemplateResponse(
            "test_result.html",
            {
                "request": request,
                "user": user,
                "test": test,
                "rows": rows,
                "total_score": score,
                "max_total": max_points,
                "show_correct": True,
            },
        )

        return templates.TemplateResponse(
            "test_run.html",
            {
                "request": request,
                "user": user,
                "test": test,
                "items": items,
                "max_points": max_points,
                "submission": submission,
                "result": result,
            },
        )


@router.get("/submissions/{submission_id}", response_class=HTMLResponse)
async def submission_detail(
    submission_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    sub = db.get(Submission, submission_id)
    if not sub:
        raise HTTPException(status_code=404, detail="submission not found")

    test = db.get(Test, sub.test_id)
    if not test:
        raise HTTPException(status_code=404, detail="test not found")
    student = db.get(User, sub.user_id) if hasattr(sub, "user_id") else None
    can_edit = user.role in ("admin", "teacher")
    if (not can_edit) and (student is None or student.id != user.id):
        raise HTTPException(status_code=403, detail="forbidden")
    show_correct = can_edit or bool(getattr(test, "show_answers_to_student", True))

    tqs: List[TestQuestion] = (
        db.query(TestQuestion)
        .filter(TestQuestion.test_id == test.id)
        .order_by(TestQuestion.order.asc())
        .all()
    )
    answers_map: dict[int, Answer] = {a.question_id: a for a in getattr(sub, "answers", [])}

    rows: List[dict] = []
    total_score = 0
    max_total = 0
    for idx, link in enumerate(tqs, 1):
        q = link.question or db.get(Question, link.question_id)
        if hasattr(q, "question") and not hasattr(q, "options"):
            if q.question is not None:
                q = q.question
        max_total += getattr(link, "points", 0) or 0

        ans = answers_map.get(q.id)
        given_raw = ""
        if ans:
            given_raw = getattr(ans, "answer_text", "") or getattr(ans, "given", "") or ""
        your_answer = given_raw or "—"
        correct_answer = ""
        recomputed_points = 0
        self_mark_allowed = (not can_edit) and student and student.id == user.id and q.answer_type == "text"

        if q.answer_type == "text":
            correct_answer = (q.correct or "") if hasattr(q, "correct") else ""
            gt = (q.correct or "").strip().lower() if hasattr(q, "correct") else ""
            uv = (given_raw or "").strip().lower()
            if gt and uv and gt == uv:
                recomputed_points = getattr(link, "points", 0) or 0
        else:
            opts = []
            if q.options:
                try:
                    opts = json.loads(q.options)
                except Exception:
                    opts = []
            try:
                correct_idx = int(q.correct) if q.correct is not None else None
            except (TypeError, ValueError):
                correct_idx = None
            try:
                user_idx = int(getattr(ans, "selected_answer_id", None)) if ans else None
            except (TypeError, ValueError):
                user_idx = None
            if correct_idx is not None and opts and 0 <= correct_idx < len(opts):
                correct_answer = str(opts[correct_idx])
            else:
                correct_answer = str(q.correct or "")
            if user_idx is not None and opts and 0 <= user_idx < len(opts):
                your_answer = str(opts[user_idx])
            if correct_idx is not None and user_idx is not None and correct_idx == user_idx:
                recomputed_points = getattr(link, "points", 0) or 0

        display_points = getattr(ans, "points", None) if ans else None
        if display_points is None:
            display_points = recomputed_points
        total_score += display_points or 0

        rows.append(
            {
                "index": idx,
                "question": q,
                "your_answer": your_answer,
                "correct_answer": correct_answer if (show_correct or self_mark_allowed) else "—",
                "score": display_points or 0,
                "max_points": getattr(link, "points", 0) or 0,
                "question_id": q.id,
                "answer_id": getattr(ans, "id", None) if ans else None,
                "self_mark_allowed": self_mark_allowed,
                "self_mark_state": bool(getattr(ans, "correct", False)) if ans else False,
            }
        )

    show_self_mark_any = any(r.get("self_mark_allowed") for r in rows)

    return templates.TemplateResponse(
        "test_result.html",
        {
            "request": request,
            "user": user,
            "test": test,
            "submission": sub,
            "student": student,
            "rows": rows,
            "total_score": total_score,
            "max_total": max_total,
            "can_edit": can_edit,
            "show_correct": show_correct,
            "show_self_mark": show_self_mark_any,
        },
    )


@router.post("/submissions/{submission_id}/set-points")
async def submission_set_points(
    submission_id: int,
    question_id: int = Form(...),
    points: int = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "teacher")),
):
    sub = db.get(Submission, submission_id)
    if not sub:
        raise HTTPException(status_code=404, detail="submission not found")

    ans = (
        db.query(Answer)
        .filter(Answer.submission_id == submission_id, Answer.question_id == question_id)
        .first()
    )
    if not ans:
        raise HTTPException(status_code=404, detail="answer not found")

    ans.points = max(points, 0)
    db.add(ans)

    new_score = (
        db.query(Answer)
        .filter(Answer.submission_id == submission_id)
        .with_entities(Answer.points)
        .all()
    )
    sub.score = sum(p[0] or 0 for p in new_score)
    db.add(sub)
    db.commit()

    return RedirectResponse(url=f"/ui/submissions/{submission_id}", status_code=303)


@router.post("/submissions/{submission_id}/self-mark", response_class=HTMLResponse)
async def submission_self_mark(
    submission_id: int,
    request: Request,
    question_id: int = Form(...),
    mark: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    sub = db.get(Submission, submission_id)
    if not sub or sub.user_id != user.id:
        raise HTTPException(status_code=404, detail="submission not found")

    link = (
        db.query(TestQuestion)
        .filter(TestQuestion.test_id == sub.test_id, TestQuestion.question_id == question_id)
        .first()
    )
    if not link:
        raise HTTPException(status_code=404, detail="question not found in test")
    q = db.get(Question, question_id)
    if not q or q.answer_type != "text":
        raise HTTPException(status_code=400, detail="self-mark allowed only for text questions")

    ans = (
        db.query(Answer)
        .filter(Answer.submission_id == sub.id, Answer.question_id == question_id)
        .first()
    )
    if not ans:
        ans = Answer(submission_id=sub.id, question_id=question_id)

    is_correct = mark == "correct"
    ans.correct = is_correct
    ans.points = getattr(link, "points", 0) if is_correct else 0
    db.add(ans)

    # пересчёт баллов попытки
    new_score = (
        db.query(Answer)
        .filter(Answer.submission_id == sub.id)
        .with_entities(Answer.points)
        .all()
    )
    sub.score = sum(p[0] or 0 for p in new_score)
    db.add(sub)
    db.commit()

    return RedirectResponse(url=f"/ui/submissions/{submission_id}", status_code=303)
